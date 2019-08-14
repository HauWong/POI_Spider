# !/usr/bin/env python
# -*- coding:utf-8 -*-

import os

import chardet
import json
import random
import openpyxl
from urllib import request, error, parse
import socket

import Transform.transform as trans


class Poi(object):

    """
    plg坐标均为WGS_84坐标
    str('lon_nw,lat_nw,lon_se,lat_se') (矩形对角点)
    或
    ('lon_1,lat_1|lon_2,lat_2|lon_3,lat_3') (多边形角点)
    """

    def __init__(self, types, plg, key, ip=None):
        self.types = parse.quote(types)
        self.plg = plg
        self.key = key
        self.poi_list = []
        self.ip = ip

    def generate_opener(self):
        if self.ip is None:
            return
        type_list = self.ip[0]
        ip_port_list = self.ip[1]
        user_agent_list = ['Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.17 (KHTML, like Gecko) Chrome/24.0.1312.57 Safari/537.17',
                           'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:6.0.2) Gecko/20100101 Firefox/6.0.2',
                           'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/28.0 Safari/537.36 OPR/15.0',
                           'Mozilla/5.0 (Windows NT 6.3; Trident/7.0; rv:11.0) like Gecko']
        ip_port_idx = random.randint(0, len(ip_port_list)-1)
        proxy = {type_list[ip_port_idx]: ip_port_list[ip_port_idx]}
        proxy_support = request.ProxyHandler(proxy)
        self.opener = request.build_opener(proxy_support)
        user_agent_idx = random.randint(0, len(user_agent_list)-1)
        self.opener.addheaders = [('User-Agent',
                              user_agent_list[user_agent_idx])]
        request.install_opener(self.opener)

    def coord_trans(self):
        if len(self.plg.split('|')) < 3:
            loc_list_old = self.plg.split(',')
            loc_nw_str_old = loc_list_old[0] + ',' + loc_list_old[1]
            loc_se_str_old = loc_list_old[2] + ',' + loc_list_old[3]
            loc_nw_str_new = trans.Transform(loc_nw_str_old).wgs_to_gcj(self.key, self.opener)
            while loc_nw_str_new is False:
                self.generate_opener()
                loc_nw_str_new = trans.Transform(loc_nw_str_old).wgs_to_gcj(self.key, self.opener)
            loc_se_str_new = trans.Transform(loc_se_str_old).wgs_to_gcj(self.key, self.opener)
            while loc_se_str_new is False:
                self.generate_opener()
                loc_se_str_new = trans.Transform(loc_se_str_old).wgs_to_gcj(self.key, self.opener)
            self.plg = loc_nw_str_new + ',' + loc_se_str_new
        else:
            point_list_old = self.plg.split('|')
            plg_str_new = ''
            for point_old in point_list_old:
                point_new = trans.Transform(point_old).wgs_to_gcj(self.key, self.opener)
                while point_new is False:
                    self.generate_opener()
                    point_new = trans.Transform(point_old).wgs_to_gcj(self.key, self.opener)
                point_str_new = '%s|' % point_new
                plg_str_new += point_str_new
            self.plg = plg_str_new[:-1]

    def gaode_urls(self):
        urls = []
        self.coord_trans()
        for page in range(1, 40):
            url = 'http://restapi.amap.com/v3/place/polygon?'\
                  'key=%s&polygon=%s&types=%s&offset=%d&page=%d&extensions=all&output=json'\
                  % (self.key, self.plg, self.types, 25, page)
            urls.append(url)
        return urls

    def gaode_search(self):
        self.generate_opener()
        uncompleted_status = True
        for url in self.gaode_urls():
            while uncompleted_status:
                try:
                    response = request.urlopen(url)
                    uncompleted_status = False
                    html = response.read()
                    response.close()
                    if chardet.detect(html)['encoding'] != 'utf-8':
                        html = html.decode('utf-8')
                    data = json.loads(html)
                    if data['status'] == '0':
                        continue
                    if data['count'] == '0':
                        break
                    print(data)
                    for item in data['pois']:
                        jname = item['name']
                        jloc = trans.Transform(item['location']).gcj_to_wgs()
                        lon, lat = jloc.split('\t')
                        jtype = item['type']
                        js_sel = [jname, float(lon), float(lat), jtype]
                        self.poi_list.append(js_sel)
                except error.URLError or socket.timeout:
                    self.generate_opener()

    def save_append(self, path, sheet_idx):
        if len(self.poi_list) == 0:
            print('No poi saved')
            return
        header = ['name', 'lon', 'lat', 'type']
        sheet_name = '%02d' % sheet_idx
        if os.path.exists(path) is False:
            wb = openpyxl.Workbook(write_only=True)
            ws = wb.create_sheet(title=sheet_name)
            ws.append(header)
        else:
            wb = openpyxl.load_workbook(path)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)
                ws.append(header)
        for poi in self.poi_list:
            ws.append(poi)
        wb.save(path)
