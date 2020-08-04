# !/usr/bin/env python
# -*- coding:utf-8 -*-

import os
import xlsxwriter
import xlrd
import openpyxl


def merge_single_poi(dir_path, name):
    write_book = xlsxwriter.Workbook(os.path.join(dir_path, name))
    work_sheet = write_book.add_worksheet()
    work_sheet.write(0, 0, 'name')
    work_sheet.write(0, 1, 'lon')
    work_sheet.write(0, 2, 'lat')
    work_sheet.write(0, 3, 'type')
    rows = 1

    for f in os.listdir(dir_path):
        if os.path.splitext(f)[1] == '.xlsx':
            filename = os.path.join(dir_path, f)
            readbook = xlrd.open_workbook(filename)
            cur_sheet = readbook.sheets()[0]
            cur_rows = cur_sheet.nrows
            for i in range(1, cur_rows):
                work_sheet.write(rows+i-1, 0, cur_sheet.row_values(i)[0])
                work_sheet.write(rows+i-1, 1, cur_sheet.row_values(i)[1])
                work_sheet.write(rows+i-1, 2, cur_sheet.row_values(i)[2])
                work_sheet.write(rows+i-1, 3, cur_sheet.row_values(i)[3])
            rows += cur_rows-1

    write_book.close()


def merge_multi_poi(dir_path):
    if not os.path.isdir(dir_path):
        raise FileNotFoundError('Wrong directory: %s' % dir_path)
    total_wb = openpyxl.Workbook(write_only=True)
    files = os.listdir(dir_path)
    for f in files:
        if f[-5:] != '.xlsx':
            continue
        poi_type = f.split('\\')[-1][:-5]
        print(poi_type)
        file_name = os.path.join(dir_path, f)
        cur_wb = openpyxl.load_workbook(file_name)
        for s in cur_wb.sheetnames:
            print('\t%s' % s)
            new_ws = total_wb.create_sheet(title='%s_%s' % (poi_type, s))
            old_ws = cur_wb[s]
            if old_ws.max_row <= 1:
                continue
            for row in old_ws.rows:
                new_ws.append(row)
    total_file_name = os.path.join(dir_path, 'total.xlsx')
    total_wb.save(total_file_name)


def merge_multi_region(dir_path):
    if not os.path.isdir(dir_path):
        raise FileNotFoundError('Wrong directory: %s' % dir_path)
    whole_wb = openpyxl.Workbook(write_only=True)
    regions = [os.path.join(dir_path, _, 'total.xlsx') for _ in os.listdir(dir_path)
               if os.path.isdir(os.path.join(dir_path, _))]
    for r in regions:
        if not os.path.exists(r):
            continue
        print(r)
        cur_wb = openpyxl.load_workbook(r)
        for s in cur_wb.sheetnames:
            cur_ws = cur_wb[s]
            if cur_ws.max_row <= 1:
                continue
            print('\t%s' % s)
            if s in whole_wb.sheetnames:
                act_ws = whole_wb[s]
            else:
                act_ws = whole_wb.create_sheet(title=s)
                act_ws.append(['name', 'lon', 'lat', 'type'])
            cur_ws.delete_rows(1)
            for row in cur_ws.rows:
                act_ws.append(row)
    whole_file_name = os.path.join(dir_path, 'whole.xlsx')
    whole_wb.save(whole_file_name)


if __name__ == '__main__':
    path = r'D:\Documents\Study\Python\Data_Model\data\FuncZ\POIs'
    # merge_multi_region(path)
    merge_multi_poi(path)
